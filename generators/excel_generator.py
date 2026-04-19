"""NexusFile AI - Excel Spreadsheet Generator"""
import os
import re
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side, numbers
from openpyxl.utils import get_column_letter
from config import GENERATED_FOLDER


# --- Color Palette ---
HEADER_FILL = PatternFill(start_color="6C5CE7", end_color="6C5CE7", fill_type="solid")
ALT_ROW_FILL = PatternFill(start_color="F8F9FA", end_color="F8F9FA", fill_type="solid")
TOTAL_FILL = PatternFill(start_color="DFE6E9", end_color="DFE6E9", fill_type="solid")
HEADER_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
DATA_FONT = Font(name="Calibri", size=11, color="2D3436")
TITLE_FONT = Font(name="Calibri", bold=True, size=16, color="6C5CE7")
TOTAL_FONT = Font(name="Calibri", bold=True, size=11, color="2D3A4A")
THIN_BORDER = Border(
    left=Side(style="thin", color="DFE6E9"),
    right=Side(style="thin", color="DFE6E9"),
    top=Side(style="thin", color="DFE6E9"),
    bottom=Side(style="thin", color="DFE6E9"),
)


def _auto_width(ws, min_width=12, max_width=40):
    """Auto-fit column widths based on content."""
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.value is not None:
                max_len = max(max_len, len(str(cell.value)))
        adjusted = min(max(max_len + 3, min_width), max_width)
        ws.column_dimensions[col_letter].width = adjusted


def generate_excel(data: dict, filename: str) -> str:
    """Generate a styled Excel workbook from structured data.

    Args:
        data: dict with 'title', 'sheets' (list of sheet dicts)
        filename: output filename (without extension)

    Returns:
        Full path to the generated .xlsx file
    """
    wb = Workbook()
    # Remove default sheet
    wb.remove(wb.active)

    sheets_data = data.get("sheets", [])
    if not sheets_data:
        sheets_data = [{"name": "Sheet1", "headers": ["Data"], "rows": [["No data"]], "formulas": []}]

    for sheet_info in sheets_data:
        sheet_name = str(sheet_info.get("name", "Sheet"))[:31]
        ws = wb.create_sheet(title=sheet_name)

        headers = sheet_info.get("headers", [])
        rows = sheet_info.get("rows", [])
        formulas = sheet_info.get("formulas", [])

        # --- Title row ---
        title = data.get("title", "Spreadsheet")
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max(len(headers), 1))
        title_cell = ws.cell(row=1, column=1, value=title)
        title_cell.font = TITLE_FONT
        title_cell.alignment = Alignment(horizontal="left", vertical="center")
        ws.row_dimensions[1].height = 35

        # Spacer row
        ws.row_dimensions[2].height = 8

        # --- Header row ---
        header_row = 3
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=header_row, column=col_idx, value=header)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = THIN_BORDER
        ws.row_dimensions[header_row].height = 28

        # Enable auto-filter
        if headers:
            ws.auto_filter.ref = f"A{header_row}:{get_column_letter(len(headers))}{header_row + len(rows)}"

        # --- Data rows ---
        for row_idx, row_data in enumerate(rows):
            excel_row = header_row + 1 + row_idx
            for col_idx, value in enumerate(row_data, 1):
                cell = ws.cell(row=excel_row, column=col_idx, value=value)
                cell.font = DATA_FONT
                cell.border = THIN_BORDER
                cell.alignment = Alignment(horizontal="left", vertical="center")

                # Alternate row colors
                if row_idx % 2 == 1:
                    cell.fill = ALT_ROW_FILL

                # Format numbers
                if isinstance(value, (int, float)):
                    cell.alignment = Alignment(horizontal="right", vertical="center")
                    if isinstance(value, float) and not value.is_integer():
                        cell.number_format = '#,##0.00'
                    elif isinstance(value, int) or (isinstance(value, float) and value.is_integer()):
                        cell.number_format = '#,##0'

        # --- Formulas ---
        if formulas and isinstance(formulas, list):
            for formula_obj in formulas:
                if not isinstance(formula_obj, dict):
                    continue
                cell_ref = formula_obj.get("cell", "")
                formula_str = formula_obj.get("formula", "")
                label = formula_obj.get("label", "")

                if cell_ref and formula_str:
                    try:
                        # Parse cell reference
                        match = re.match(r"([A-Z]+)(\d+)", cell_ref.upper())
                        if match:
                            cell = ws[cell_ref.upper()]
                            cell.value = formula_str
                            cell.font = TOTAL_FONT
                            cell.fill = TOTAL_FILL
                            cell.border = THIN_BORDER
                            cell.alignment = Alignment(horizontal="right", vertical="center")

                            # Add label in previous column if exists
                            if label and cell.column > 1:
                                label_cell = ws.cell(row=cell.row, column=cell.column - 1)
                                if not label_cell.value:
                                    label_cell.value = label
                                    label_cell.font = TOTAL_FONT
                                    label_cell.fill = TOTAL_FILL
                                    label_cell.border = THIN_BORDER
                    except Exception:
                        pass

        # Freeze panes (freeze header row)
        ws.freeze_panes = f"A{header_row + 1}"

        # Auto-fit columns
        _auto_width(ws)

    # ===== SAVE =====
    safe_filename = "".join(c if c.isalnum() or c in (' ', '-', '_') else '_' for c in filename)
    output_path = os.path.join(GENERATED_FOLDER, f"{safe_filename}.xlsx")
    wb.save(output_path)
    return output_path
